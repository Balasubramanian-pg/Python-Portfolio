from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.formatting.rule import DataBarRule
from datetime import date

# -------------------------
# CONFIG
# -------------------------
FILE_NAME = "Habit_Dashboard.xlsx"
FONT = "Tenorite"  # will fall back if not installed

NAVY = "1F2A44"
PURPLE = "7C3AED"
CYAN = "38BDF8"
LIGHT_BG = "F8FAFC"
GRAY = "E5E7EB"

TODAY = date.today().strftime("%A, %b %d")

HABITS = [
    ("Drinking Water", 0.5),
    ("Praying", 0.75),
    ("Medicine", 0.0),
    ("Exercise", 0.0),
    ("Reading", 1.0),
    ("Meditation", 0.5)
]

# -------------------------
# WORKBOOK
# -------------------------
wb = Workbook()
ws = wb.active
ws.title = "Dashboard"

ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 18

# -------------------------
# LEFT NAV PANEL
# -------------------------
for r in range(1, 40):
    cell = ws.cell(r, 1)
    cell.fill = PatternFill("solid", fgColor=NAVY)

ws["A2"].value = "Tempo"
ws["A2"].font = Font(name=FONT, size=16, bold=True, color="FFFFFF")
ws["A2"].alignment = Alignment(horizontal="center")

nav_items = ["Dashboard", "To-do", "Habits", "Insights", "Streaks", "Settings"]
row = 5
for item in nav_items:
    ws[f"A{row}"].value = item
    ws[f"A{row}"].font = Font(name=FONT, size=11, color="FFFFFF")
    ws[f"A{row}"].alignment = Alignment(horizontal="left")
    row += 2

# -------------------------
# HEADER
# -------------------------
ws.merge_cells("B2:F2")
ws["B2"].value = TODAY
ws["B2"].font = Font(name=FONT, size=14, bold=True)
ws["B2"].alignment = Alignment(horizontal="left")

# -------------------------
# KPI TILES
# -------------------------
kpis = [
    ("Completed", "10%"),
    ("On progress", "45%"),
    ("Pending", "45%"),
]

col = 2
for title, value in kpis:
    ws.merge_cells(start_row=4, start_column=col, end_row=6, end_column=col)
    c = ws.cell(4, col)
    c.value = title
    c.font = Font(name=FONT, size=11)
    ws.cell(5, col).value = value
    ws.cell(5, col).font = Font(name=FONT, size=18, bold=True)
    ws.cell(4, col).alignment = Alignment(horizontal="center")
    ws.cell(5, col).alignment = Alignment(horizontal="center")
    ws.cell(4, col).fill = PatternFill("solid", fgColor=LIGHT_BG)
    ws.cell(5, col).fill = PatternFill("solid", fgColor=LIGHT_BG)
    col += 1

# -------------------------
# TODAY'S TASKS
# -------------------------
ws["B8"].value = "Today's tasks"
ws["B8"].font = Font(name=FONT, size=13, bold=True)

start_row = 10
for habit, progress in HABITS:
    ws[f"B{start_row}"].value = habit
    ws[f"B{start_row}"].font = Font(name=FONT, size=11)

    ws[f"C{start_row}"].value = progress
    ws[f"C{start_row}"].number_format = "0%"

    start_row += 1

bar = DataBarRule(
    start_type="num", start_value=0,
    end_type="num", end_value=1,
    color=PURPLE
)

ws.conditional_formatting.add(f"C10:C{start_row-1}", bar)

# -------------------------
# STREAK DONUT CHART
# -------------------------
ws["E8"].value = "Streaks"
ws["E8"].font = Font(name=FONT, size=13, bold=True)

ws["E10"].value = "Completed"
ws["E11"].value = "Missed"
ws["F10"].value = 64
ws["F11"].value = 36

chart = DoughnutChart()
chart.title = "This Week"

data = Reference(ws, min_col=6, min_row=10, max_row=11)
labels = Reference(ws, min_col=5, min_row=10, max_row=11)
chart.add_data(data, titles_from_data=False)
chart.set_categories(labels)
chart.width = 8
chart.height = 8

ws.add_chart(chart, "E13")

# -------------------------
# SAVE
# -------------------------
wb.save(FILE_NAME)

print("Dashboard created:", FILE_NAME)
